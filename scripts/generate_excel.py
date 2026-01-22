#!/usr/bin/env python3
"""
Enhanced Excel generation with data cleaning and auto-numbering.
"""
import sys
import json
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict, Any


def clean_text(text: str) -> str:
    """
    Clean OCR artifacts and noise from text.

    Args:
        text: Raw text from OCR

    Returns:
        Cleaned text
    """
    if not text or not isinstance(text, str):
        return ""

    # Remove common OCR noise characters
    text = text.replace('|', '').replace('||', '')
    text = text.replace('，', '').replace('。', '')
    text = text.replace('"', '').replace('"', '')

    # Remove leading/trailing punctuation and whitespace
    text = text.strip()
    text = re.sub(r'^[,，.:：;；\s]+', '', text)
    text = re.sub(r'[,，.:：;；\s]+$', '', text)

    # Remove multiple spaces
    text = re.sub(r'\s+', ' ', text)

    # Fix common OCR errors in Chinese
    text = text.replace('基浮', '悬浮')  # Common OCR mistake
    text = text.replace('折又', '折叠')
    text = text.replace('钦件', '软件')
    text = text.replace('重法', '方法')

    return text.strip()


def validate_and_clean_data(data: Dict[str, str]) -> Dict[str, str]:
    """
    Validate and clean copyright data.

    Args:
        data: Raw copyright data dictionary

    Returns:
        Cleaned data dictionary
    """
    cleaned = {}

    # Clean all text fields
    for key in ['著作权人', '软件名称', '首次发表日期', '权利取得方式', '权利范围']:
        value = data.get(key, '')
        cleaned[key] = clean_text(value)

    # Handle registration number (登记号) - keep numbers only
    reg_num = clean_text(data.get('登记号', ''))
    # Validate format: YYYYSRxxxxxxx
    if reg_num and re.match(r'^\d{4}SR\d+$', reg_num):
        cleaned['登记号'] = reg_num
    else:
        cleaned['登记号'] = reg_num  # Keep as is, user can review

    # Ignore the original serial number from OCR - we'll auto-generate it
    # But save it for reference if needed
    original_serial = clean_text(data.get('序号', ''))
    cleaned['_原始序号'] = original_serial if original_serial else ''

    return cleaned


def generate_excel(output_file: Path, data_list: List[Dict[str, str]]):
    """
    Generate Excel file from list of copyright data with enhanced formatting.

    Args:
        output_file: Path to output Excel file
        data_list: List of copyright data dictionaries
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "软件著作权清单"

    # Define headers
    headers = ["序号", "著作权人", "软件名称", "首次发表日期", "权利取得方式", "权利范围", "登记号", "备注"]

    # Style configuration
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    data_font = Font(name='微软雅黑', size=10)
    centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = thin_border

    # Clean and validate data
    cleaned_data = []
    for item in data_list:
        cleaned = validate_and_clean_data(item)

        # Skip if no key fields found (likely OCR garbage)
        if not any([
            cleaned.get('软件名称'),
            cleaned.get('著作权人'),
            cleaned.get('登记号')
        ]):
            continue

        cleaned_data.append(cleaned)

    # Write data with auto-numbering
    for row_idx, item in enumerate(cleaned_data, 2):
        # Auto-generate sequential number
        auto_number = row_idx - 1

        # Prepare row data
        row_data = [
            auto_number,  # Auto-generated sequential number
            item.get("著作权人", ""),
            item.get("软件名称", ""),
            item.get("首次发表日期", ""),
            item.get("权利取得方式", ""),
            item.get("权利范围", ""),
            item.get("登记号", ""),
            f"原序号: {item.get('_原始序号', '')}" if item.get('_原始序号') else ""
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # Alignment: center for number/date, left for text
            if col_idx in [1, 4, 7]:  # 序号, 首次发表日期, 登记号
                cell.alignment = centered_alignment
            else:
                cell.alignment = left_alignment

    # Set column widths
    column_widths = {
        'A': 8,   # 序号
        'B': 30,  # 著作权人
        'C': 50,  # 软件名称
        'D': 18,  # 首次发表日期
        'E': 15,  # 权利取得方式
        'F': 12,  # 权利范围
        'G': 20,  # 登记号
        'H': 20   # 备注
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze first row
    ws.freeze_panes = 'A2'

    # Set row height
    ws.row_dimensions[1].height = 25

    # Save file
    try:
        wb.save(output_file)
        print(f"✅ Excel文件已生成: {output_file}")
        print(f"📊 总计 {len(cleaned_data)} 条有效记录")

        # Automatically open the file on macOS
        import subprocess
        try:
            subprocess.run(["open", str(output_file)], check=True)
            print(f"📂 正在打开文件...")
        except Exception:
            pass

    except Exception as e:
        print(f"❌ 保存Excel文件出错: {e}", file=sys.stderr)
        raise


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_excel.py <output_file.xlsx> <data.json>")
        sys.exit(1)

    output_file = Path(sys.argv[1])
    input_data_path = Path(sys.argv[2])

    try:
        with open(input_data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Ensure data is a list
        if isinstance(data, dict):
            data = [data]

        generate_excel(output_file, data)

    except Exception as e:
        print(f"❌ 错误: {e}", file=sys.stderr)
        sys.exit(1)
